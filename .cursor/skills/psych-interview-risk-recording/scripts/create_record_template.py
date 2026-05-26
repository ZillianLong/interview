#!/usr/bin/env python3
"""Generate per-candidate interview record Excel (12 columns × 8 dimensions)."""

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

ROOT = Path(__file__).resolve().parents[4]
RECORD_DIR = ROOT / "效果验证" / "面试记录"

HEADERS = [
    "维度",
    "问题",
    "候选人回答",
    "追问",
    "候选人回答",
    "考察点1分析",
    "考察点2分析",
    "风险等级",
    "等级描述",
    "候选人自我报告",
    "候选人对题反馈",
    "判断不一致说明",
]

# 与 面试观察.md 定稿同步
DIMENSIONS = [
    {
        "dim": "抑郁悲观",
        "question": "在学习、实习或是日常工作里，有没有哪件事做完之后自己特别满意，过后回想起来也依旧觉得做得很好？麻烦讲一讲这次经历，说说当时具体是什么情形。",
        "followup": "做完之后，心里是什么感受？",
    },
    {
        "dim": "焦虑不安",
        "question": "在日常的学习和工作中，大家都有压力比较大的时候。提到压力，你觉得最近让你感到特别有压力的事情是什么？拿一次具体的经历讲讲。",
        "followup": "碰到那种情况，你一般会怎么应对？",
    },
    {
        "dim": "冲动违规",
        "question": "平时上学、实习或者上班期间，或者日常生活里，你觉得哪些规则或要求不太合理，甚至让你感到压抑？",
        "followup": "你为什么觉得这些规矩不太合理？有没有你认为更好的做法？",
    },
    {
        "dim": "行为古怪",
        "question": "结合你过往的经历聊聊，和同学或者同事一起商量事情时，他们的反应与你的预期相差很大的情况，当时具体的情形是什么？",
        "followup": "你自己后来怎么看这件事？",
    },
    {
        "dim": "偏执多疑",
        "question": "在与同事或者同学相处时，难免遇到当时觉得没什么，但是事后越想越不舒服的情况。根据你的实际情况，聊聊相关的经历以及你当时的想法与感受吧。",
        "followup": "后来你们聊开了吗？你对最后的结果怎么看的？",
    },
    {
        "dim": "喜怒无常",
        "question": "请说说最近让你特别来气，而且你平复了好久才消气的事情？简单讲讲整件事的来龙去脉。",
        "followup": "你当时是怎么反应的？在这件事情上大概与对方僵持了多久？",
    },
    {
        "dim": "刻板固执",
        "question": "学习或者工作中，有时候需要和大家分工配合。把一些事情交给别人做的时候，拿一次经历具体分享下具体的经过吧。",
        "followup": "团队合作中，别人的做法与你的做法不一致时，你一般是怎么处理的？",
    },
    {
        "dim": "妄想离奇",
        "question": "全程自然观察",
        "followup": "全程自然观察",
    },
]


def create_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "面试记录"

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row_idx, item in enumerate(DIMENSIONS, start=2):
        ws.cell(row=row_idx, column=1, value=item["dim"])
        ws.cell(row=row_idx, column=2, value=item["question"])
        ws.cell(row=row_idx, column=4, value=item["followup"])
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    widths = [12, 32, 26, 22, 26, 28, 28, 8, 28, 26, 22, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for row_idx in range(2, 10):
        ws.row_dimensions[row_idx].height = 80

    return wb


def main():
    parser = argparse.ArgumentParser(description="Create interview record Excel template")
    parser.add_argument(
        "--id",
        default="候选人",
        help="候选人标识，用于文件名 面试记录_{id}.xlsx",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite existing file",
    )
    args = parser.parse_args()

    RECORD_DIR.mkdir(parents=True, exist_ok=True)
    out_path = RECORD_DIR / f"面试记录_{args.id}.xlsx"

    if out_path.exists() and not args.force:
        print(f"Refusing to overwrite (use --force): {out_path}", file=sys.stderr)
        sys.exit(1)

    wb = create_workbook()
    wb.save(out_path)
    print(f"Saved {out_path}")


if __name__ == "__main__":
    main()
