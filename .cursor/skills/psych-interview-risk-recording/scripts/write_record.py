#!/usr/bin/env python3
"""Validate filled interview record Excel."""

import argparse
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from score_descriptions import get_description, load_from_xlsx

EXPECTED_DIMS = [
    "抑郁悲观",
    "焦虑不安",
    "冲动违规",
    "行为古怪",
    "偏执多疑",
    "喜怒无常",
    "刻板固执",
    "妄想离奇",
]

EXPECTED_HEADERS = [
    "维度",
    "问题",
    "候选人回答",
    "追问",
    "候选人回答",
    "考察点1分析",
    "考察点2分析",
    "风险等级",
    "等级描述",
    "候选人自我报告",
    "候选人对题反馈",
    "判断不一致说明",
]

VALID_LEVELS = {"低", "中", "高", ""}
COL_RISK = 8
COL_DESC = 9
COL_SELF = 10
COL_Q_FB = 11
COL_MISMATCH = 12


def validate(path: Path) -> list[str]:
    errors: list[str] = []
    scores = load_from_xlsx()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, len(EXPECTED_HEADERS) + 1)]
    if headers != EXPECTED_HEADERS:
        errors.append(f"表头应为 {EXPECTED_HEADERS}，实际为 {headers}")

    dims_found = []
    for row in range(2, 2 + len(EXPECTED_DIMS)):
        dim = ws.cell(row, 1).value
        if dim:
            dims_found.append(dim)
        level = (ws.cell(row, COL_RISK).value or "").strip()

        if level not in VALID_LEVELS:
            errors.append(f"行 {row} [{dim}] 风险等级非法: {level!r}（须为 低/中/高）")

        if level in {"中", "高"}:
            a1 = (ws.cell(row, 6).value or "").strip()
            a2 = (ws.cell(row, 7).value or "").strip()
            if not a1 and not a2:
                errors.append(f"行 {row} [{dim}] 标「{level}」但考察点分析为空")

        desc = (ws.cell(row, COL_DESC).value or "").strip()
        if level in {"低", "中", "高"} and dim:
            expected = scores.get((dim, level), get_description(dim, level))
            if expected and desc and expected not in desc and desc not in expected:
                errors.append(
                    f"行 {row} [{dim}] 等级描述与 维度得分描述.xlsx 可能不一致"
                )

        mismatch = (ws.cell(row, COL_MISMATCH).value or "").strip()
        self_rpt = (ws.cell(row, COL_SELF).value or "").strip()
        if mismatch and not self_rpt:
            errors.append(f"行 {row} [{dim}] 有不一致说明但自我报告为空")

    if dims_found != EXPECTED_DIMS:
        errors.append(f"维度顺序/名称应为 {EXPECTED_DIMS}，实际为 {dims_found}")

    wb.close()
    return errors


def main():
    parser = argparse.ArgumentParser(description="Validate interview record Excel")
    parser.add_argument("path", type=Path, help="面试记录 xlsx 路径")
    args = parser.parse_args()

    if not args.path.exists():
        print(f"File not found: {args.path}", file=sys.stderr)
        sys.exit(1)

    errors = validate(args.path)
    if errors:
        print("Validation FAILED:")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    print(f"OK: {args.path}")


if __name__ == "__main__":
    main()
